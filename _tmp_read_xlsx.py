import openpyxl
import sys
import json

path = r"C:\Users\איתי\OneDrive\שולחן העבודה\מוזמנים חתן + כלה + הורי החתן .xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

result = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    result[sheet_name] = {
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "rows": rows,
    }

with open(r"c:\Users\איתי\OneDrive\שולחן העבודה\פרויקטים\Events-Mannagement\Events-Mannagement\_tmp_xlsx_output.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("DONE", len(result), "sheets")
